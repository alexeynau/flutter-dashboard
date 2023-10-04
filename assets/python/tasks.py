from datetime import datetime
from bs4 import BeautifulSoup
from RPA.Browser.Selenium import Selenium
from RPA.Excel.Files import Files as Excel
from openpyxl import load_workbook
from robocorp.tasks import task
from RPA.Desktop.OperatingSystem import OperatingSystem
from collections import defaultdict
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import pandas as pd
import numpy as np
import logging
import os
import json
from dotenv import load_dotenv

load_dotenv()

try:
    os.remove('parse_log.log')
except OSError:
    pass

logging.basicConfig(filename='parse_log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',
                    encoding='utf-8')

output_directory = os.environ.get("ROBOT_ARTIFACTS")
shared_directory = os.path.join(output_directory, "shared")
app1_file_path = os.environ.get("APP1_FILE_PATH")
app2_file_path = os.environ.get("APP2_FILE_PATH")
print("APP1_FILE_PATH", app1_file_path, "APP2_FILE_PATH", app2_file_path)


@task
def web_preprocessor():
    from first_table import pars_usd, customs_duties_parser, parcing_first_table, parsing_brent_cost, urals_parser, second_table
    print("Start parser")
    logging.info("Start parser")
    pars_usd.usd_kurs()
    pars_excel(app1_file_path)
    urals_parser.start(app1_file_path)
    customs_duties_parser.start(app1_file_path)
    parcing_first_table.start(app1_file_path)
    parsing_brent_cost.start(app1_file_path)
    second_table.to_application_1(app1_file_path, app2_file_path)

# TODO: implement in postprocessor


def pars_excel(file_path):
    from first_table.parcing_first_table import set_cell_value

    # Загрузите данные из monthly_exchange_rate.xlsx
    monthly_data = pd.read_excel('monthly_exchange_rate.xlsx').T

    months = {'январь': 1, 'февраль': 2, 'март': 3, 'апрель': 4, 'май': 5,
              'июнь': 6, 'июль': 7, 'август': 8, 'сентябрь': 9, 'октябрь': 10,
              'ноябрь': 11, 'декабрь': 12}

    logging.info(monthly_data)

    # Загрузите данные из Приложение_1.xlsx
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Анализ_БК+ББ']
    counter = 4
    shipment_date = sheet[f"M{counter}"].value
    reciving_date = sheet[f"P{counter}"].value
    while shipment_date is not None:
        logging.info(type(shipment_date))
        if isinstance(shipment_date, datetime):
            logging.info(f"shipment_date: {shipment_date.month}")
            course = monthly_data[shipment_date.month].loc['Средний курс']
        else:
            course = monthly_data[months[shipment_date.lower(
            ).replace(" ", "")]].loc['Средний курс']

        logging.info(f"shipment_date course: {course}")
        set_cell_value(f'Y{counter}', course, sheet)
        if isinstance(reciving_date, datetime):
            logging.info(f"reciving_date: {reciving_date.month}")
            course = monthly_data[reciving_date.month].loc['Средний курс']
        else:
            course = monthly_data[months[reciving_date.lower(
            ).replace(" ", "")]].loc['Средний курс']
        logging.info(f"reciving_date course: {course}")
        set_cell_value(f'Z{counter}', course, sheet)
        counter += 1
        shipment_date = sheet[f"M{counter}"].value
        reciving_date = sheet[f"P{counter}"].value
    wb.save(file_path)
    wb.close()


def after_usd_kurs():
    #  Загрузите данные из quarterly_exchange_rate.xlsx
    quarterly_data = pd.read_excel('quarterly_exchange_rate.xlsx')

    # Загрузите данные из Приложение_1.xlsx
    app1_data = pd.read_excel(app1_file_path)

    # Задайте столбцы кварталов, с которыми нужно сопоставить данные
    quarters_to_match = ["1 кв", "2 кв", "3 кв", "4 кв"]

    # Обновите данные в Приложение_1.xlsx
    for quarter in quarters_to_match:
        app1_data[quarter] = quarterly_data["Средний курс"]

    # Сохраните обновленные данные в Приложение_1.xlsx
    app1_data.to_excel(app1_file_path, index=False)


def add_to_json(df, title, plot_type, chats_data, hidden=None):
    data = []
    for column in list(df.columns):
        value = {
            "name": column[1],
            "series": list(df[column])
        }
        data.append(value)
    plot_name = {
        "plot_name": title,
        "x": list(df['x'].columns)[0],
        "y": list(df.xs('y', axis=1, level=0))
    }
    if hidden is not None:
        plot_name['hidden'] = [list(df.xs('hidden', axis=1, level=0))[:2],
            list(df.xs('hidden', axis=1, level=0))[2:]]

    chats_data["data"] = chats_data["data"] + data
    chats_data["charts"][plot_type].append(plot_name)
    return chats_data


def is_numeric(val):
    try:
        float(val)
        return True
    except ValueError:
        return False


def parse_dollar_ex_rate():
    exchange_rate_df = pd.read_excel('monthly_exchange_rate.xlsx')
    exchange_rate_df['Месяц'] = exchange_rate_df['Месяц'].replace({1: 'Январь',
                                                                   2: 'Февраль',
                                                                   3: 'Март',
                                                                   4: 'Апрель',
                                                                   5: 'Май',
                                                                   6: 'Июнь',
                                                                   7: 'Июль',
                                                                   8: 'Август',
                                                                   9: 'Сентябрь',
                                                                   10: 'Октябрь',
                                                                   11: 'Ноябрь',
                                                                   12: 'Декабрь'})
    exchange_rate_df['Средний курс'] = exchange_rate_df['Средний курс'].apply(
        lambda x: round(x, 2))
    columns = pd.MultiIndex.from_tuples([('x', col) if col != 'Средний курс' else (
        'y', 'Средний курс') for col in exchange_rate_df.columns])
    exchange_rate_df.columns = columns
    return exchange_rate_df


def open_excel():
    excel = Excel()
    excel.open_workbook(app1_file_path, data_only=True)
    excel.set_active_worksheet('Анализ_БК+ББ')
    rows = excel.read_worksheet_as_table()
    return rows

def get_company(rows):
    counter = 1
    row = rows.get_row(counter)
    black_list = ['Переоценка остатков ДС', 'Продажа валюты',
                  'Переоценка ДЗ и КЗ на конец каждого периода']
    companies_list = ['Компания 1', 'Company ABC',
                      'A-Нефтегаз', 'Компания ААА']
    companies = defaultdict(dict)

    try:
        while True:
            if (row['A'] is None) and (row['Y'] is None) and (row['Z'] is None):
                company_name = row['J']
                if (company_name not in companies.keys()) and (company_name not in black_list) and (
                        rows.get_row(counter + 1)['J'] is not None) and (company_name in companies_list):
                    companies[company_name]['row_index'] = counter + 1
            counter += 1
            row = rows.get_row(counter)
    except IndexError:
        print(f'Компании закончились')
    return companies


def get_merged_cells():
    wb = load_workbook(app1_file_path)
    sheet = wb['Анализ_БК+ББ']
    merged_cells_values_start = []
    merged_cells_values_end = []

    for merged_range in sheet.merged_cells.ranges:
        merged_range_str = str(merged_range)
        merged_cells_values_start.append(merged_range_str.split(':')[0])
        merged_cells_values_end.append(merged_range_str.split(':')[-1])

    return merged_cells_values_start, merged_cells_values_end


def add_column_to_list(companies, merged_cells_values_start, merged_cells_values_end, rows, column_name, name, month=False, date=False):
    for company_elem in companies.keys():
        company_row_index = companies[company_elem]['row_index']
        added_column = rows.get_column(column_name)
        columns_list = []
        for index, item in enumerate(list(added_column.values())[company_row_index:]):

            if (f'{column_name}{company_row_index+index}' in merged_cells_values_start):
                item = rows.get_cell(
                    (company_row_index + index) - 1, column_name)

            if (item is None):
                break

            if (item in ('#DIV/0!', '#REF!')):
                columns_list.append(0)
                continue

            if month and isinstance(item, str):
                columns_list.append(item)
            elif month:
                columns_list.append(item.strftime('%B'))
            elif date:
                columns_list.append(item.strftime('%Y-%m-%d'))
            else:
                columns_list.append(item)

        index = next((i for i, x in enumerate(
            reversed(columns_list)) if x != 0), len(columns_list))
        columns_list = columns_list[:-index or None]
        companies[company_elem][name] = columns_list

    return companies


def get_processed_df(company, company_name, x: str, y: list, hidden: list = None, reverse: bool = False, agg_col=None):
    columns = [x] + y + (hidden if hidden is not None else [])
    min_length = min([len(company[item]) for item in columns])
    data = {('x', x): company[x][:min_length]}
    for col in y:
        if ('y', col) not in data:
            data[('y', col)] = []
        data[('y', col)].extend(company[col][:min_length])

    if hidden:
        for col in hidden:
            if ('hidden', col) not in data:
                data[('hidden', col)] = []
            data[('hidden', col)].extend(
                company[col][:min_length])

    df = pd.DataFrame(data)
    new_columns = pd.MultiIndex.from_tuples(
        [(col[0], f'{col[1]}_{company_name}') for col in df.columns])
    df.columns = new_columns
    if agg_col is not None:
        name = f"{agg_col}_{company_name}"
        df = df.groupby(('x', name), as_index=False).mean()
    if reverse:
        df[('y', f"Эффект_{company_name}")] = df[(
            'y', f"Эффект_{company_name}")].apply(lambda x: None)
        df = df.transpose().reset_index()
        df.drop(columns=['level_0'], inplace=True)
        new_columns = pd.MultiIndex.from_tuples(
            [('x', f'Макропараметры_{company_name}') if col == 'level_1' else ('y', f'сделка_{col+1}_{company_name}') for col in df.columns])
        df.columns = new_columns
        num_columns = len(df.columns) - 1
        new_column_names = [
            f'сделка_{i}' for i in range(1, num_columns + 1)]
        df = df.rename(columns={col: new_name for col, new_name in zip(
            df.columns[1:], new_column_names)})
    return df


def process_chart_to_json(companies, chats_data, x, y, title, plot_type, hidden=None, reverse=False, agg_col=None):
    for company in list(companies.keys())[:2]:
        df = get_processed_df(
            companies[company], company, x, y, hidden, reverse, agg_col)
        chats_data = add_to_json(
            df, f'{title}_{company}', plot_type, chats_data, hidden)
    return chats_data


@task
def after_update_postprocessor():
    chats_data = {"data": [], "charts": defaultdict(list)}
    rows = open_excel()
    companies = get_company(rows)
    merged_cells_values_start, merged_cells_values_end = get_merged_cells()

    # Курс доллара по месяцам
    exchange_rate_df = parse_dollar_ex_rate()
    chats_data = add_to_json(
        exchange_rate_df, "Курс доллара по месяцам", "plots", chats_data=chats_data)

    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'M', 'Дата отгрузки', month=False, date=True)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'P', 'Дата поступления выручки', month=False, date=True)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'Y', 'курс на дату отгрузки', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'Z', 'курс на дату поступления', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AD', 'Freight,$/барр.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AV', 'Freight ,млн руб.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AC', 'Spread,$/барр.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AL', 'Spread,млн руб.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AB', 'Цена Brent,$/барр.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AK', 'Цена Brent,млн руб.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AJ', 'реализация (отгрузка), рубли', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AR', 'поступление, рубли', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AI', 'реализация,USD', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'AS', 'Курсовые разницы между отгрузкой и поступлением денежных ср-в, руб', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'Z', 'курс на дату поступления ден средств (погашения ДЗ)', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'L', 'Месяц отгрузки', month=True)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'O', 'Месяц поступления выручки', month=True)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'CZ', 'Коэффициент перевода барр./т.', month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DB', "Допдоход к рынку (к цене НДПИ Platts) млн руб", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DA', "Допдоход к рынку (к цене НДПИ Argus) млн руб", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DC', "Зачисление процентов,млн руб", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DD', "Макропараметры (brent/ discount/ escalation) млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DE', "Макропараметры (spread) млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DF', "Freight, млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DG', "Коммерческие за искл. Freight, млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DH', "Экспортная пошлина, млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DI', "Курсовые разницы (отгрузка / оплата)млн руб", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DJ', "откл НДПИ (макропараметры)", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DK', "откл НДПИ уплач от расч НДПИ (курс оплаты)млн руб", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DL', "откл НДПИ (Argus) и НДПИ (Plats Urals Med), млн руб.", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'DM', "Эффект", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'K', "Условия поставок", month=False)
    companies = add_column_to_list(
        companies, merged_cells_values_start, merged_cells_values_end, rows, 'CS', "Сумма денежной выручки, млн руб.", month=False)

    charts = [{'x': 'Дата отгрузки', 'y': ['Freight,$/барр.', 'Freight ,млн руб.'], 'title': "Курс Freight по дате отгрузки",
               'plot_type': 'plots', 'hidden': None, 'reverse': False, 'agg_col': None},
              {'x': 'Дата отгрузки', 'y': ['Spread,$/барр.', 'Spread,млн руб.'], 'title': "Spread на дату отгрузки",
               'plot_type': 'plots', 'hidden': None, 'reverse': False, 'agg_col': None},
              {'x': 'Дата отгрузки', 'y': ['Цена Brent,$/барр.', 'Цена Brent,млн руб.'], 'title': "Brent на дату отгрузки",
               'plot_type': 'plots', 'hidden': None, 'reverse': False, 'agg_col': None},
              {'x': 'Условия поставок', 'y': ['Сумма денежной выручки, млн руб.'], 'title': "Средняя выручка по условиям договора",
               'plot_type': 'bar_chart', 'hidden': None, 'reverse': False, 'agg_col': 'Условия поставок'},
              {'x': 'Допдоход к рынку (к цене НДПИ Platts) млн руб', 'y': ['Допдоход к рынку (к цене НДПИ Argus) млн руб',
                                                                           'Зачисление процентов,млн руб',
                                                                           'Макропараметры (brent/ discount/ escalation) млн руб.',
                                                                           'Макропараметры (spread) млн руб.',
                                                                           'Freight, млн руб.',
                                                                           'Коммерческие за искл. Freight, млн руб.',
                                                                           'Экспортная пошлина, млн руб.',
                                                                           'Курсовые разницы (отгрузка / оплата)млн руб',
                                                                           'откл НДПИ (макропараметры)',
                                                                           'откл НДПИ уплач от расч НДПИ (курс оплаты)млн руб',
                                                                           'откл НДПИ (Argus) и НДПИ (Plats Urals Med), млн руб.',
                                                                           'Эффект'],
               'title': "Отклонения в макропараметрах",
               'plot_type': 'waterfall', 'hidden': None, 'reverse': True, 'agg_col': None},
              {'x': 'Месяц отгрузки', 'y': ['Коэффициент перевода барр./т.'], 'title': "Коэффициент перевода барр./т. в месяц отгрузки",
               'plot_type': 'bar_chart', 'hidden': None, 'reverse': False, 'agg_col': None},
              {'x': 'Месяц поступления выручки', 'y': ['Коэффициент перевода барр./т.'], 'title': "Коэффициент перевода барр./т. в месяц поступления выручки",
               'plot_type': 'bar_chart', 'hidden': None, 'reverse': False, 'agg_col': None},
              {'x': 'реализация,USD', 'y': ['реализация (отгрузка), рубли', 'поступление, рубли'], 'title': "Разница в итоговой прибыли при вариации сроков реализации этапов",
               'plot_type': 'plots', 'hidden': ['курс на дату отгрузки', 'Дата отгрузки', 'курс на дату поступления', 'Дата поступления выручки'], 'reverse': False, 'agg_col': None}]
    for chart in charts:
        chats_data = process_chart_to_json(
            companies, chats_data, chart['x'], chart['y'], chart['title'], chart['plot_type'], chart['hidden'], chart['reverse'], chart['agg_col'])
    with open(os.path.join(shared_directory, 'workitems.json'), "w", encoding='utf-8') as outfile:
        json.dump(chats_data, outfile, ensure_ascii=False)
